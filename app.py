import streamlit as st
import pandas as pd
import yfinance as yf
import numpy as np
import io
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, ScatterChart, Reference, Series
from openpyxl.chart.trendline import Trendline
from openpyxl.formatting.rule import ColorScaleRule

# ==========================================
# 1. HELPER FUNCTIONS
# ==========================================

def format_indian_symbol(raw_input: str) -> str:
    symbol = raw_input.strip().upper()
    index_map = {
    "NIFTY": "^NSEI", "NIFTY 50": "^NSEI", "NIFTY50": "^NSEI",
    "SENSEX": "^BSESN", "BSE SENSEX": "^BSESN",
    "BANKNIFTY": "^NSEBANK", "NIFTY BANK": "^NSEBANK",
    "NIFTY 500": "^CRSLDX", "NIFTY500": "^CRSLDX",
    "NIFTY MIDCAP": "^NSEMDCP50", "NIFTY MIDCAP 50": "^NSEMDCP50",
    "NIFTY SMALLCAP": "^NSI200", "NIFTY SMALLCAP 200": "^NSI200",
    "NIFTY IT": "^CNXIT", "NIFTY AUTO": "^CNXAUTO", "NIFTY FMCG": "^CNXFMCG",
    "NIFTY PHARMA": "^CNXPHARMA", "NIFTY METAL": "^CNXMETAL", "NIFTY REALTY": "^CNXREALTY",
    "NIFTY ENERGY": "^CNXENERGY", "NIFTY PSU BANK": "^CNXPSUBANK",
    "NIFTY PRIVATE BANK": "^NIFTYPVTBANK", "NIFTY FINANCIAL SERVICES": "^CNXFINSERVICE",
    "NIFTY MEDIA": "^CNXMEDIA", "NIFTY CONSUMPTION": "^CNXCONSUMPTION",
    "NIFTY COMMODITIES": "^CNXCOMMODITIES", "NIFTY INFRASTRUCTURE": "^CNXINFRA",
    "NIFTY HEALTHCARE": "^CNXHEALTHCARE", "NIFTY OIL & GAS": "^CNXOILANDGAS",
    "AUTO": "^CNXAUTO", "IT": "^CNXIT", "FMCG": "^CNXFMCG", "PHARMA": "^CNXPHARMA",
    "METAL": "^CNXMETAL", "REALTY": "^CNXREALTY", "BANK": "^NSEBANK",
    }
    if symbol in index_map: return index_map[symbol]
    if symbol.startswith("^") or symbol.endswith(".NS") or symbol.endswith(".BO"):
        return symbol
    return f"{symbol}.NS"

def clean_sheet_name(name: str) -> str:
    return name.replace("^", "").replace(":", "_").replace("/", "_").strip()[:30].upper()

@st.cache_data(show_spinner=False)
def get_financial_metrics(ticker_symbol: str, de_type: str) -> tuple[float, float, float]:
    de_ratio, tax_rate, market_cap = 0.0, 0.25, 0.0 
    try:
        t = yf.Ticker(ticker_symbol)
        info = t.info or {}
        
        market_cap_raw = info.get("marketCap")
        if market_cap_raw is not None:
            market_cap = float(market_cap_raw)

        if de_type == "Market Value D/E":
            total_debt = info.get("totalDebt")
            if total_debt is None:
                try:
                    bs = t.balance_sheet
                    if not bs.empty and "Total Debt" in bs.index:
                        total_debt = bs.loc["Total Debt"].iloc[0]
                except: pass
                
            if total_debt is not None and market_cap > 0:
                de_ratio = float(total_debt / market_cap)
        else:
            d_e_raw = info.get("debtToEquity")
            if d_e_raw is not None and isinstance(d_e_raw, (int, float)):
                de_ratio = (d_e_raw / 100.0 if d_e_raw > 5 else float(d_e_raw))
            else:
                try:
                    bs = t.balance_sheet
                    if not bs.empty and "Total Debt" in bs.index and "Stockholders Equity" in bs.index:
                        total_debt = bs.loc["Total Debt"].iloc[0]
                        total_equity = bs.loc["Stockholders Equity"].iloc[0]
                        if total_equity and total_equity != 0:
                            de_ratio = float(total_debt / total_equity)
                except: pass

        try:
            inc = t.income_stmt
            if not inc.empty and "Tax Provision" in inc.index and "Pretax Income" in inc.index:
                tax_provision = inc.loc["Tax Provision"].iloc[0]
                pretax_income = inc.loc["Pretax Income"].iloc[0]
                if pretax_income > 0 and tax_provision > 0:
                    calc_tax = float(tax_provision / pretax_income)
                    if 0.05 <= calc_tax <= 0.45: tax_rate = calc_tax
        except: pass
    except: pass
    return round(de_ratio, 4), round(tax_rate, 4), market_cap

# ==========================================
# 2. EXCEL GENERATION LOGIC (WITH CHARTS)
# ==========================================

def generate_market_excel(
    data_dict: dict,
    financials_dict: dict,
    main_assets: list,
    comp_map: dict,
    benchmark_symbol: str,
    start_date: str,
    end_date: str,
    dash_df: pd.DataFrame,
    industry_df: pd.DataFrame,
    weight_table_df: pd.DataFrame,
    calc_breakdown_df: pd.DataFrame,
    peer_detail_data: list
) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    NAVY_FILL = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    ZEBRA_FILL = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    FONT_TITLE = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    FONT_BOLD = Font(name="Calibri", size=10, bold=True)
    FONT_METRIC = Font(name="Calibri", size=11, bold=True, color="1B365D")

    BORDER_THIN = Border(
        left=Side(style="thin", color="D5D8DC"), right=Side(style="thin", color="D5D8DC"),
        top=Side(style="thin", color="D5D8DC"), bottom=Side(style="thin", color="D5D8DC"),
    )
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

    all_symbols = list(data_dict.keys())
    ordered_sheets = [s for s in main_assets if s in all_symbols]
    ordered_sheets += [s for s in all_symbols if s not in main_assets and s != benchmark_symbol]
    if benchmark_symbol in all_symbols:
        ordered_sheets.append(benchmark_symbol)

    for user_symbol in ordered_sheets:
        sheet_title = clean_sheet_name(user_symbol)
        if sheet_title not in wb.sheetnames:
            wb.create_sheet(title=sheet_title)

    bench_clean = clean_sheet_name(benchmark_symbol)

    for user_symbol in ordered_sheets:
        df = data_dict[user_symbol]
        sheet_title = clean_sheet_name(user_symbol)
        ws = wb[sheet_title]
        de_ratio, tax_rate, mc = financials_dict.get(user_symbol, (0.0, 0.25, 0.0))

        ws.merge_cells("A1:G1")
        ws["A1"] = f"MARKET DATA: {sheet_title}"
        ws["A1"].font = FONT_TITLE
        ws["A1"].fill = NAVY_FILL

        headers = ["Date", "Open", "High", "Low", "Close", "Volume", "Daily Return"]
        for c_i, h_text in enumerate(headers, start=1):
            c = ws.cell(row=3, column=c_i, value=h_text)
            c.fill, c.font, c.alignment, c.border = HEADER_FILL, FONT_HEADER, ALIGN_CENTER, BORDER_THIN

        start_row = 4
        for i, (idx_dt, row_data) in enumerate(df.iterrows()):
            r_idx = start_row + i
            ws.cell(row=r_idx, column=1, value=idx_dt.strftime("%Y-%m-%d")).alignment = ALIGN_CENTER
            for col, val in enumerate(["Open", "High", "Low", "Close"], start=2):
                ws.cell(row=r_idx, column=col, value=float(row_data[val])).number_format = "₹#,##0.00"
            ws.cell(row=r_idx, column=6, value=int(row_data["Volume"])).number_format = "#,##0"
            
            if i == 0: ws.cell(row=r_idx, column=7, value=0.0).number_format = "0.00%"
            else: ws.cell(row=r_idx, column=7, value=f"=(E{r_idx}-E{r_idx-1})/E{r_idx-1}").number_format = "0.00%"
            
            for c in range(1, 8):
                ws.cell(row=r_idx, column=c).border = BORDER_THIN
                if i % 2 == 1: ws.cell(row=r_idx, column=c).fill = ZEBRA_FILL

        end_row = start_row + len(df) - 1

        metrics = [
            ("Avg Daily Return", f"=AVERAGE(G4:G{end_row})", "0.000%"),
            ("Annual Return", "=J4*252", "0.00%"),
            ("Daily Vol", f"=_xlfn.STDEV.S(G4:G{end_row})", "0.000%"),
            ("Annual Vol", "=J6*SQRT(252)", "0.00%"),
            ("Daily Var", f"=_xlfn.VAR.S(G4:G{end_row})", "0.00000"),
            (f"Covariance ({bench_clean})", f"=_xlfn.COVARIANCE.S(G4:G{end_row}, '{bench_clean}'!G4:G{end_row})", "0.00000"),
            (f"Bench Var", f"=_xlfn.VAR.S('{bench_clean}'!G4:G{end_row})", "0.00000"),
            ("Beta (Ratio)", "=IF(J10=0, 1, J9/J10)", "0.00"),
            ("Levered Beta", f"=SLOPE(G4:G{end_row}, '{bench_clean}'!G4:G{end_row})", "0.00"),
            ("D/E Ratio", de_ratio, "0.00%"),
            ("Tax Rate", tax_rate, "0.00%"),
            ("Unlevered Beta", "=J12/(1+(1-J14)*J13)", "0.00"),
            ("Correlation", f"=CORREL(G4:G{end_row}, '{bench_clean}'!G4:G{end_row})", "0.000")
        ]
        
        for m_idx, (label, form, fmt) in enumerate(metrics, start=4):
            ws.cell(row=m_idx, column=9, value=label).font = FONT_BOLD
            c_val = ws.cell(row=m_idx, column=10, value=form)
            c_val.font = FONT_METRIC
            c_val.number_format = fmt

        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["I"].width = 25
        ws.column_dimensions["J"].width = 15

        if user_symbol in main_assets and benchmark_symbol in data_dict:
            sc = ScatterChart()
            sc.title = f"{sheet_title} vs {bench_clean} Regression"
            sc.style = 13
            sc.x_axis.title = f"{bench_clean} Returns"
            sc.y_axis.title = f"{sheet_title} Returns"
            sc.width, sc.height = 15, 10
            
            bench_sheet = wb[bench_clean]
            x_values = Reference(bench_sheet, min_col=7, min_row=4, max_row=end_row)
            y_values = Reference(ws, min_col=7, min_row=4, max_row=end_row)
            
            series = Series(y_values, x_values, title_from_data=False)
            series.marker.symbol = "circle"
            series.graphicalProperties.line.noFill = True
            series.trendline = Trendline(trendlineType="linear")
            
            sc.series.append(series)
            ws.add_chart(sc, "M15")

    # ==========================================
    # Executive Summary Sheet
    # ==========================================
    ws_summary = wb.create_sheet(title="Executive Summary", index=0)
    ws_summary.views.sheetView[0].showGridLines = True
    ws_summary.merge_cells("A1:K1")
    ws_summary["A1"] = f"QUANTITATIVE ANALYSIS ({start_date} to {end_date})"
    ws_summary["A1"].font, ws_summary["A1"].fill = FONT_TITLE, NAVY_FILL

    headers_summary = ["Asset", "Avg Daily Ret", "Ann Return", "Ann Volatility", "Levered Beta", "D/E Ratio", "Tax Rate", "Unlevered Beta", "Peer Unlevered Beta", "Peer Relevered Beta", "Correlation"]
    
    for col_idx, text in enumerate(headers_summary, start=1):
        c = ws_summary.cell(row=3, column=col_idx, value=text)
        c.fill, c.font, c.alignment, c.border = HEADER_FILL, FONT_HEADER, ALIGN_CENTER, BORDER_THIN
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = 20

    for idx, row in dash_df.iterrows():
        s_row = 4 + idx
        ws_summary.cell(row=s_row, column=1, value=row["Asset"]).font = FONT_BOLD
        ws_summary.cell(row=s_row, column=2, value=row["Ann Return"]/252).number_format = "0.000%"
        ws_summary.cell(row=s_row, column=3, value=row["Ann Return"]).number_format = "0.00%"
        ws_summary.cell(row=s_row, column=4, value=row["Ann Volatility"]).number_format = "0.00%"
        ws_summary.cell(row=s_row, column=5, value=row["Raw Levered Beta"]).number_format = "0.00"
        ws_summary.cell(row=s_row, column=6, value=row["D/E Ratio"]).number_format = "0.00%"
        ws_summary.cell(row=s_row, column=7, value=financials_dict.get(row["Asset"], (0.0, 0.25, 0.0))[1]).number_format = "0.00%"
        ws_summary.cell(row=s_row, column=8, value=row["Unlevered Beta"]).number_format = "0.00"
        ws_summary.cell(row=s_row, column=9, value=row["Peer Avg Unlevered Beta"]).number_format = "0.00"
        ws_summary.cell(row=s_row, column=10, value=row["Peer Relevered Beta"]).number_format = "0.00"
        ws_summary.cell(row=s_row, column=11, value=row["Correlation"]).number_format = "0.000"
        for c_i in range(1, 12):
            ws_summary.cell(row=s_row, column=c_i).border = BORDER_THIN

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Beta Comparison by Asset"
    chart.y_axis.title = "Beta Value"
    chart.x_axis.title = "Assets"

    cats_ref = Reference(ws_summary, min_col=1, min_row=4, max_row=3 + len(dash_df))
    s1 = Series(Reference(ws_summary, min_col=5, min_row=3, max_row=3 + len(dash_df)), title_from_data=True)
    s2 = Series(Reference(ws_summary, min_col=8, min_row=3, max_row=3 + len(dash_df)), title_from_data=True)
    s3 = Series(Reference(ws_summary, min_col=10, min_row=3, max_row=3 + len(dash_df)), title_from_data=True)
    
    chart.series.extend([s1, s2, s3])
    chart.set_categories(cats_ref)
    chart.width = 18
    chart.height = 10
    ws_summary.add_chart(chart, "M3")

    # ==========================================
    # Industry Scenarios Sheet
    # ==========================================
    ws_ind = wb.create_sheet(title="Industry Scenarios", index=1)
    ws_ind.views.sheetView[0].showGridLines = True
    
    # Table 1: Target Relevering Scenarios
    ws_ind.merge_cells("A1:G1")
    ws_ind["A1"] = "INDUSTRY BETA SCENARIOS (RELEVERED BY TARGET)"
    ws_ind["A1"].font, ws_ind["A1"].fill = FONT_TITLE, NAVY_FILL

    headers_ind = [
        "Target Asset", "Calc 1: Global Simple Unlevered", "Calc 1: Target Relevered", 
        "Calc 2: Global Weighted Unlevered", "Calc 2: Target Relevered", 
        "Calc 3: Peer Weighted Unlevered", "Calc 3: Target Relevered"
    ]
    for col_idx, text in enumerate(headers_ind, start=1):
        c = ws_ind.cell(row=3, column=col_idx, value=text)
        c.fill, c.font, c.alignment, c.border = HEADER_FILL, FONT_HEADER, ALIGN_CENTER, BORDER_THIN
        ws_ind.column_dimensions[get_column_letter(col_idx)].width = 25

    for idx, row in industry_df.iterrows():
        s_row = 4 + idx
        ws_ind.cell(row=s_row, column=1, value=row["Target Asset"]).font = FONT_BOLD
        ws_ind.cell(row=s_row, column=2, value=row["Calc 1: Global Simple Unlevered"]).number_format = "0.00"
        ws_ind.cell(row=s_row, column=3, value=row["Calc 1: Target Relevered"]).number_format = "0.00"
        ws_ind.cell(row=s_row, column=4, value=row["Calc 2: Global Weighted Unlevered"]).number_format = "0.00"
        ws_ind.cell(row=s_row, column=5, value=row["Calc 2: Target Relevered"]).number_format = "0.00"
        ws_ind.cell(row=s_row, column=6, value=row["Calc 3: Peer Weighted Unlevered"]).number_format = "0.00"
        ws_ind.cell(row=s_row, column=7, value=row["Calc 3: Target Relevered"]).number_format = "0.00"
        
        for c_i in range(1, 8):
            ws_ind.cell(row=s_row, column=c_i).border = BORDER_THIN

    # Table 2: Market Cap, Weights & Weighted Average Breakdown
    t2_start_row = 5 + len(industry_df) + 2
    ws_ind.merge_cells(f"A{t2_start_row-1}:D{t2_start_row-1}")
    ws_ind[f"A{t2_start_row-1}"] = "MARKET CAP & WEIGHTED AVERAGE CALCULATIONS (ALL COMPANIES)"
    ws_ind[f"A{t2_start_row-1}"].font, ws_ind[f"A{t2_start_row-1}"].fill = FONT_TITLE, NAVY_FILL

    headers_weight = ["Company / Asset", "Unlevered Beta", "Market Cap", "Industry Weight (%)"]
    for col_idx, text in enumerate(headers_weight, start=1):
        c = ws_ind.cell(row=t2_start_row, column=col_idx, value=text)
        c.fill, c.font, c.alignment, c.border = HEADER_FILL, FONT_HEADER, ALIGN_CENTER, BORDER_THIN

    for idx, row in weight_table_df.iterrows():
        r_idx = t2_start_row + 1 + idx
        ws_ind.cell(row=r_idx, column=1, value=row["Company"]).font = FONT_BOLD
        ws_ind.cell(row=r_idx, column=2, value=row["Unlevered Beta"]).number_format = "0.00"
        ws_ind.cell(row=r_idx, column=3, value=row["Market Cap"]).number_format = "#,##0"
        ws_ind.cell(row=r_idx, column=4, value=row["Industry Weight (%)"]).number_format = "0.00%"
        for c_i in range(1, 5):
            ws_ind.cell(row=r_idx, column=c_i).border = BORDER_THIN

    # Table 3: Detailed Calculation Breakdown for Unlevered Betas
    t3_start_row = t2_start_row + len(weight_table_df) + 3
    ws_ind.merge_cells(f"A{t3_start_row-1}:C{t3_start_row-1}")
    ws_ind[f"A{t3_start_row-1}"] = "UNLEVERED BETA CALCULATION BREAKDOWN (BY METHOD)"
    ws_ind[f"A{t3_start_row-1}"].font, ws_ind[f"A{t3_start_row-1}"].fill = FONT_TITLE, NAVY_FILL

    headers_calc = ["Method", "Resulting Unlevered Beta", "Basis"]
    for col_idx, text in enumerate(headers_calc, start=1):
        c = ws_ind.cell(row=t3_start_row, column=col_idx, value=text)
        c.fill, c.font, c.alignment, c.border = HEADER_FILL, FONT_HEADER, ALIGN_CENTER, BORDER_THIN

    for idx, row in calc_breakdown_df.iterrows():
        r_idx = t3_start_row + 1 + idx
        ws_ind.cell(row=r_idx, column=1, value=row["Method"]).font = FONT_BOLD
        ws_ind.cell(row=r_idx, column=2, value=row["Resulting Unlevered Beta"]).number_format = "0.00"
        ws_ind.cell(row=r_idx, column=3, value=row["Basis"])
        for c_i in range(1, 4):
            ws_ind.cell(row=r_idx, column=c_i).border = BORDER_THIN

    # Table 4: Peer Weighted Unlevered Beta (Calc 3) - Per Target Detail
    t4_start_row = t3_start_row + len(calc_breakdown_df) + 3
    ws_ind.merge_cells(f"A{t4_start_row-1}:E{t4_start_row-1}")
    ws_ind[f"A{t4_start_row-1}"] = "PEER WEIGHTED UNLEVERED BETA (CALC 3) - PER TARGET DETAIL"
    ws_ind[f"A{t4_start_row-1}"].font, ws_ind[f"A{t4_start_row-1}"].fill = FONT_TITLE, NAVY_FILL

    headers_peer_detail = ["Target Asset", "Peer Company", "Peer Unlevered Beta", "Peer Market Cap", "Peer Weight in Group (%)"]
    for col_idx, text in enumerate(headers_peer_detail, start=1):
        c = ws_ind.cell(row=t4_start_row, column=col_idx, value=text)
        c.fill, c.font, c.alignment, c.border = HEADER_FILL, FONT_HEADER, ALIGN_CENTER, BORDER_THIN

    for idx, row in peer_detail_data.iterrows():
        r_idx = t4_start_row + 1 + idx
        ws_ind.cell(row=r_idx, column=1, value=row["Target Asset"]).font = FONT_BOLD
        ws_ind.cell(row=r_idx, column=2, value=row["Peer Company"])
        ws_ind.cell(row=r_idx, column=3, value=row["Peer Unlevered Beta"]).number_format = "0.00"
        ws_ind.cell(row=r_idx, column=4, value=row["Peer Market Cap"]).number_format = "#,##0"
        ws_ind.cell(row=r_idx, column=5, value=row["Peer Weight in Group (%)"]).number_format = "0.00%"
        for c_i in range(1, 6):
            ws_ind.cell(row=r_idx, column=c_i).border = BORDER_THIN

    # ==========================================
    # Correlation Heatmap Sheet
    # ==========================================
    ws_corr = wb.create_sheet(title="Correlation Heatmap", index=2)
    returns_df = pd.DataFrame()
    for sym in main_assets:
        if sym in data_dict:
            returns_df[sym] = data_dict[sym]['Close'].pct_change().fillna(0.0)
    corr_matrix = returns_df.corr()
    
    ws_corr.cell(row=1, column=1, value="Asset").font = FONT_BOLD
    for c_idx, col_name in enumerate(corr_matrix.columns, 2):
        ws_corr.cell(row=1, column=c_idx, value=col_name).font = FONT_BOLD
        
    for r_idx, row_name in enumerate(corr_matrix.index, 2):
        ws_corr.cell(row=r_idx, column=1, value=row_name).font = FONT_BOLD
        for c_idx, col_name in enumerate(corr_matrix.columns, 2):
            val = corr_matrix.loc[row_name, col_name]
            c = ws_corr.cell(row=r_idx, column=c_idx, value=val)
            c.number_format = "0.00"
            c.border = BORDER_THIN
            
    rule = ColorScaleRule(start_type='min', start_color='F8696B', 
                          mid_type='num', mid_value=0, mid_color='FFEB84', 
                          end_type='max', end_color='63BE7B')
    ws_corr.conditional_formatting.add(f"B2:{get_column_letter(len(corr_matrix.columns)+1)}{len(corr_matrix.index)+1}", rule)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ==========================================
# 3. STREAMLIT APP UI & LOGIC
# ==========================================

st.set_page_config(page_title="Beta & Risk Analytics", layout="wide", initial_sidebar_state="expanded")

st.title("📈 Advanced Cost of Capital & Beta Dashboard")
st.markdown("Analyze market risk, compute unlevered betas, and handle custom competitor peer groups dynamically.")

st.sidebar.header("Configuration")

mode = st.sidebar.radio(
    "Select Output Mode",
    ["Interactive Dashboard Only", "Excel Workbook with Charts Only", "Generate Both"]
)

start_date = st.sidebar.date_input("Start Date", datetime(2021, 4, 1))
end_date = st.sidebar.date_input("End Date", datetime(2026, 3, 31))
benchmark_input = st.sidebar.text_input("Benchmark Index", "NIFTY 50")

de_ratio_type = st.sidebar.radio(
    "D/E Ratio Type",
    ["Book Value D/E", "Market Value D/E"],
    help="Choose whether to calculate the Debt-to-Equity ratio using the Book Value of Equity or the current Market Capitalization."
)

with st.sidebar.expander("⚙️ Optional Competitor Mapping (Advanced)", expanded=False):
    st.markdown("""
    *Specify custom peers per company to override default peer averaging.*  
    **Format:** `Target = Peer1, Peer2`  
    **Example:**  
    `TCS = INFY, WIPRO`  
    `RELIANCE = ONGC, BPCL`  
    *(If left blank, defaults to averaging other Target Companies).*
    """)
    custom_peers_str = st.text_area("Competitor Rules", height=120, placeholder="TCS = INFY, WIPRO")

main_tickers_str = st.text_input("Target Companies (comma-separated)", "TVSMOTOR, HEROMOTOCO, EICHERMOT")

if st.button("Run Financial Analysis", type="primary"):
    main_assets = [s.strip().upper() for s in main_tickers_str.split(",") if s.strip()]
    if not main_assets:
        st.error("Please enter at least one target company.")
        st.stop()

    comp_map = {}
    additional_peers = []
    if custom_peers_str.strip():
        for line in custom_peers_str.strip().split("\n"):
            if "=" in line:
                target, peers = line.split("=")
                target = target.strip().upper()
                peer_list = [p.strip().upper() for p in peers.split(",") if p.strip()]
                comp_map[target] = peer_list
                additional_peers.extend(peer_list)

    all_symbols_to_fetch = list(set(main_assets + additional_peers))
    bench_symbol = benchmark_input.strip().upper()
    if bench_symbol not in all_symbols_to_fetch:
        all_symbols_to_fetch.append(bench_symbol)

    with st.spinner("Downloading market data and computing metrics..."):
        fetched_data = {}
        financials_data = {}
        
        yf_start = start_date.strftime("%Y-%m-%d")
        yf_end = end_date.strftime("%Y-%m-%d")

        for sym in all_symbols_to_fetch:
            ticker = format_indian_symbol(sym)
            try:
                stock = yf.Ticker(ticker)
                df = stock.history(start=yf_start, end=yf_end)
                if not df.empty:
                    df.index = df.index.tz_localize(None)
                    fetched_data[sym] = df.round(2)
                    financials_data[sym] = get_financial_metrics(ticker, de_ratio_type)
            except Exception as e:
                st.warning(f"Failed to fetch {sym}: {e}")

        if bench_symbol not in fetched_data:
            st.error(f"Could not fetch benchmark '{bench_symbol}'. Please verify the ticker.")
            st.stop()

        common_dates = set(fetched_data[bench_symbol].index)
        for sym in fetched_data:
            common_dates = common_dates.intersection(set(fetched_data[sym].index))
        common_dates = sorted(list(common_dates))

        for sym in list(fetched_data.keys()):
            fetched_data[sym] = fetched_data[sym].loc[common_dates].copy()

        bench_df = fetched_data[bench_symbol]['Close']
        bench_ret = bench_df.pct_change().fillna(0.0)

        metrics_cache = {}
        for sym in all_symbols_to_fetch:
            if sym == bench_symbol: continue
            if sym not in fetched_data: continue
            target_ret = fetched_data[sym]['Close'].pct_change().fillna(0.0)
            
            aligned = pd.concat([target_ret, bench_ret], axis=1)
            aligned.columns = ['Target', 'Bench']
            
            mean_ret = aligned['Target'].mean()
            ann_ret = mean_ret * 252
            ann_vol = aligned['Target'].std(ddof=1) * np.sqrt(252)
            
            bench_var = aligned['Bench'].var(ddof=1)
            cov = aligned['Target'].cov(aligned['Bench'])
            
            beta_l = cov / bench_var if bench_var != 0 else 1.0
            de, tax, mc = financials_data.get(sym, (0.0, 0.25, 0.0))
            beta_u = beta_l / (1 + (1 - tax) * de)
            corr = aligned['Target'].corr(aligned['Bench'])
            
            metrics_cache[sym] = {
                "mean_ret": mean_ret, "ann_ret": ann_ret, "ann_vol": ann_vol,
                "beta_l": beta_l, "de": de, "tax": tax, "beta_u": beta_u, "corr": corr, "mc": mc
            }

        valid_companies = [s for s in all_symbols_to_fetch if s in metrics_cache and s != bench_symbol]
        total_mc = 0.0
        if valid_companies:
            global_simple_u_beta = np.mean([metrics_cache[c]["beta_u"] for c in valid_companies])
            total_mc = sum(metrics_cache[c]["mc"] for c in valid_companies)
            if total_mc > 0:
                global_mc_u_beta = sum(metrics_cache[c]["beta_u"] * metrics_cache[c]["mc"] for c in valid_companies) / total_mc
            else:
                global_mc_u_beta = global_simple_u_beta 
        else:
            global_simple_u_beta, global_mc_u_beta = 1.0, 1.0

        weight_table_data = []
        for c in valid_companies:
            c_mc = metrics_cache[c]["mc"]
            c_weight = (c_mc / total_mc) if total_mc > 0 else 0.0
            weight_table_data.append({
                "Company": c,
                "Unlevered Beta": metrics_cache[c]["beta_u"],
                "Market Cap": c_mc,
                "Industry Weight (%)": c_weight
            })
        weight_table_df = pd.DataFrame(weight_table_data)

        industry_scenario_data = []
        dashboard_data = []
        calc_breakdown_data = []
        peer_detail_rows = []

        for sym in main_assets:
            if sym not in metrics_cache: continue
            m = metrics_cache[sym]
            de, tax = m["de"], m["tax"]
            
            if sym in comp_map and comp_map[sym]:
                peers = [p for p in comp_map[sym] if p in metrics_cache]
            else:
                peers = [p for p in all_symbols_to_fetch if p in metrics_cache and p != bench_symbol]
            
            peer_u_beta = np.mean([metrics_cache[p]["beta_u"] for p in peers]) if peers else np.nan
            peer_rel_beta = peer_u_beta * (1 + (1 - m["tax"]) * m["de"]) if not np.isnan(peer_u_beta) else np.nan

            dashboard_data.append({
                "Asset": sym,
                "Ann Return": m["ann_ret"],
                "Ann Volatility": m["ann_vol"],
                "Raw Levered Beta": m["beta_l"],
                "D/E Ratio": m["de"],
                "Unlevered Beta": m["beta_u"],
                "Peer Avg Unlevered Beta": peer_u_beta,
                "Peer Relevered Beta": peer_rel_beta,
                "Correlation": m["corr"]
            })

            c1_rel = global_simple_u_beta * (1 + (1 - tax) * de)
            c2_rel = global_mc_u_beta * (1 + (1 - tax) * de)

            c3_unlev = np.nan
            c3_rel = np.nan
            if peers:
                peer_total_mc = sum(metrics_cache[p]["mc"] for p in peers)
                if peer_total_mc > 0:
                    c3_unlev = sum(metrics_cache[p]["beta_u"] * metrics_cache[p]["mc"] for p in peers) / peer_total_mc
                    for p in peers:
                        p_weight = metrics_cache[p]["mc"] / peer_total_mc
                        peer_detail_rows.append({
                            "Target Asset": sym,
                            "Peer Company": p,
                            "Peer Unlevered Beta": metrics_cache[p]["beta_u"],
                            "Peer Market Cap": metrics_cache[p]["mc"],
                            "Peer Weight in Group (%)": p_weight
                        })
                else:
                    c3_unlev = peer_u_beta 
                    for p in peers:
                        peer_detail_rows.append({
                            "Target Asset": sym,
                            "Peer Company": p,
                            "Peer Unlevered Beta": metrics_cache[p]["beta_u"],
                            "Peer Market Cap": metrics_cache[p]["mc"],
                            "Peer Weight in Group (%)": 1.0 / len(peers)
                        })
                c3_rel = c3_unlev * (1 + (1 - tax) * de)

            industry_scenario_data.append({
                "Target Asset": sym,
                "Calc 1: Global Simple Unlevered": global_simple_u_beta,
                "Calc 1: Target Relevered": c1_rel,
                "Calc 2: Global Weighted Unlevered": global_mc_u_beta,
                "Calc 2: Target Relevered": c2_rel,
                "Calc 3: Peer Weighted Unlevered": c3_unlev,
                "Calc 3: Target Relevered": c3_rel
            })

        calc_breakdown_data.append({
            "Method": "Global Simple Unlevered Beta (Calc 1)",
            "Resulting Unlevered Beta": global_simple_u_beta,
            "Basis": "Simple average across all companies (targets + peers)"
        })
        calc_breakdown_data.append({
            "Method": "Global Weighted Unlevered Beta (Calc 2)",
            "Resulting Unlevered Beta": global_mc_u_beta,
            "Basis": "Market-cap weighted average across all companies"
        })

    dash_df = pd.DataFrame(dashboard_data)
    industry_df = pd.DataFrame(industry_scenario_data)
    calc_breakdown_df = pd.DataFrame(calc_breakdown_data)
    peer_detail_df = pd.DataFrame(peer_detail_rows)

    st.session_state['dash_df'] = dash_df
    st.session_state['industry_df'] = industry_df
    st.session_state['weight_table_df'] = weight_table_df
    st.session_state['calc_breakdown_df'] = calc_breakdown_df
    st.session_state['peer_detail_df'] = peer_detail_df
    st.session_state['fetched_data'] = fetched_data
    st.session_state['financials_data'] = financials_data
    st.session_state['main_assets'] = main_assets
    st.session_state['comp_map'] = comp_map
    st.session_state['bench_symbol'] = bench_symbol
    st.session_state['yf_start'] = yf_start
    st.session_state['yf_end'] = yf_end
    st.session_state['ran_analysis'] = True

if 'ran_analysis' in st.session_state and st.session_state['ran_analysis']:
    dash_df = st.session_state['dash_df']
    industry_df = st.session_state['industry_df']
    weight_table_df = st.session_state['weight_table_df']
    calc_breakdown_df = st.session_state['calc_breakdown_df']
    peer_detail_df = st.session_state['peer_detail_df']
    fetched_data = st.session_state['fetched_data']
    financials_data = st.session_state['financials_data']
    main_assets = st.session_state['main_assets']
    comp_map = st.session_state['comp_map']
    bench_symbol = st.session_state['bench_symbol']
    yf_start = st.session_state['yf_start']
    yf_end = st.session_state['yf_end']

    show_dash = mode in ["Interactive Dashboard Only", "Generate Both"]
    show_excel = mode in ["Excel Workbook with Charts Only", "Generate Both"]

    if show_dash:
        st.divider()
        tab1, tab2, tab3 = st.tabs(["📊 Main Dashboard", "🌐 Industry Beta Scenarios", "📈 Charts & Regressions"])
        
        with tab1:
            st.subheader("Executive Summary Metrics")
            format_dict = {
                "Ann Return": "{:.2%}", "Ann Volatility": "{:.2%}", "D/E Ratio": "{:.2%}",
                "Raw Levered Beta": "{:.2f}", "Unlevered Beta": "{:.2f}",
                "Peer Avg Unlevered Beta": "{:.2f}", "Peer Relevered Beta": "{:.2f}", "Correlation": "{:.3f}"
            }
            st.dataframe(dash_df.style.format(format_dict).background_gradient(cmap='Blues', subset=['Raw Levered Beta', 'Peer Relevered Beta']), use_container_width=True)

        with tab2:
            st.subheader("Target Relevering based on Industry Averages")
            st.dataframe(industry_df.style.format("{:.2f}", subset=['Calc 1: Global Simple Unlevered', 'Calc 1: Target Relevered', 'Calc 2: Global Weighted Unlevered', 'Calc 2: Target Relevered', 'Calc 3: Peer Weighted Unlevered', 'Calc 3: Target Relevered']), use_container_width=True)

            st.markdown("### Market Cap & Weighted Average Calculations (All Companies)")
            st.dataframe(weight_table_df.style.format({"Unlevered Beta": "{:.2f}", "Market Cap": "{:,.0f}", "Industry Weight (%)": "{:.2%}"}), use_container_width=True)

            st.markdown("### Unlevered Beta Calculation Breakdown (By Method)")
            st.dataframe(calc_breakdown_df.style.format({"Resulting Unlevered Beta": "{:.3f}"}), use_container_width=True)

            st.markdown("### Peer Weighted Unlevered Beta (Calc 3) - Per Target Detail")
            st.dataframe(peer_detail_df.style.format({"Peer Unlevered Beta": "{:.3f}", "Peer Market Cap": "{:,.0f}", "Peer Weight in Group (%)": "{:.2%}"}), use_container_width=True)

        with tab3:
            st.subheader("Visual Regressions")
            for sym in main_assets:
                if sym in fetched_data:
                    fig = px.scatter(
                        x=fetched_data[bench_symbol]['Close'].pct_change().dropna(),
                        y=fetched_data[sym]['Close'].pct_change().dropna(),
                        trendline="ols",
                        labels={"x": f"{bench_symbol} Return", "y": f"{sym} Return"},
                        title=f"{sym} vs {bench_symbol} Regression"
                    )
                    st.plotly_chart(fig, use_container_width=True)

    if show_excel:
        st.divider()
        st.subheader("📥 Download Excel Report")
        excel_bytes = generate_market_excel(
            fetched_data, financials_data, main_assets, comp_map, benchmark_symbol, 
            yf_start, yf_end, dash_df, industry_df, weight_table_df, calc_breakdown_df, peer_detail_df
        )
        st.download_button(
            label="Download Complete Excel Workbook (.xlsx)",
            data=excel_bytes,
            file_name=f"Beta_Cost_Of_Capital_Analysis_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
